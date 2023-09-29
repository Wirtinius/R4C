from openpyxl import Workbook
from datetime import date, timedelta
from .models import Robot
import os
from django.conf import settings

def create_excel():
  today = date.today()
  start = today - timedelta(days=today.weekday())
  end = start + timedelta(days=6)
  wb = Workbook()
  wb.remove_sheet(wb["Sheet"])
  num = 0
  for model in Robot.objects.values_list('model').distinct():
    wb.create_sheet(f"{model[0]}")
    ws = wb[wb.sheetnames[num]]
    ws.append(["Модель", "Версия", "Количество за неделю"])
    for robot in Robot.objects.filter(model=model[0]):
      if start <= robot.created.date() <= end:
        if robot.version not in [item[1].value for item in ws]:
          ws.append([robot.model, robot.version, Robot.objects.filter(model=robot.model, version=robot.version).count()])
    num += 1
  file_path = os.path.join(settings.MEDIA_ROOT, 'robot_production_report.xlsx')
  os.makedirs(os.path.dirname(file_path), exist_ok=True)
  wb.save(file_path)

  return file_path
